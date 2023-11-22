#!/usr/bin/env python3
"""
    Python script to read ciderSDR.xml file and build a BOM
    in Excel format.
    
    Peter Fetterer <KB3GTN@GMAIL.COM>
"""

import os
import sys
import xmltodict
import openpyxl

def check_bom(parts_database):
    """ look at parts in database and see if each part as a 
        "manufacture" and "manufacture part number" entries. 
    """
    components_result = True
    for comp in parts_database:
        refdes = comp['@ref']
        value = comp['value']
        manufacturer = None
        partnumber = None
        try:
            fields = comp['fields']        
            for field in comp['fields']['field']:
                if field['@name'] == 'Manufacturer':
                    manufacturer = field['#text']
                if field['@name'] == 'Manufacturer Part Number':
                    partnumber = field['#text']
        except:           
            pass
        
        if manufacturer == None or partnumber == None:
            if refdes[0] == "R" or refdes[0] == "C" or refdes[0] == "D" or refdes == "J":
                print(f"Component '{refdes}' is missing 'Manufacturer' and/or 'Manufacture Part Number' entry.  Treating as generic {refdes[0]} component..") 
            else:
                print(f"Component '{refdes}' is missing 'Manufacturer' and/or 'Manufacturer Part Number'")
                components_result = False
    
    return components_result
    

def group_bom_items(parts_database):
    """ return a list of components that are group by footprint / manufactuer / part number """
    # list of list of refdes groups
    line_items = []
    list_of_refdes_records = {} # dict of refdes : part parameters to be processed
    
    # build list of part dictonaries for each refdes in part database
    for part in parts_database:
        list_of_refdes_records[part['@ref']] = part

    # build line_items lists of parts grouped by value,manufacture,partnumber
    for key,dvalue in list_of_refdes_records.items():
        # get part parameters
        refdes = key
        value = dvalue['value']
        manufacturer = None
        partnumber = None
        try:
            fields = dvalue['fields']        
            for field in dvalue['fields']['field']:
                if field['@name'] == 'Manufacturer':
                    manufacturer = field['#text']
                if field['@name'] == 'Manufacturer Part Number':
                    partnumber = field['#text']
        except KeyError as e:
            print(f"Generic Part infered for {refdes}")
            if refdes[0] == "R" or refdes[0] == "C" or refdes[0] == "D" or refdes[0] == "J":
                Manufacturer = "Generic"
                partnumber = value
            else:    
                print("Caught exception: {e.str()} when looking for manufactuer / manufactuer part numbers in parts database..")
                print("Parts Database missing required keys..")
                return -1
        
        # see if we have a line_item with a matching part number
        found_li = False
        for li in line_items:
            if li['Manufacturer Part Number'] == partnumber:
                # found potental line item
                if li['value'] == value and li['manufacturer'] == manufacturer:
                    # Looks the same.. Add to line item
                    li['refdes_list'].append(refdes)
                    found_li = True
                    
        if found_li == False:
            # If we didn't find a match in our list of line items.. add new line item
            new_line_item = { 'refdes_list' : [refdes], 'value' : value, 'manufacturer' : manufacturer, 'Manufacturer Part Number' : partnumber }
            line_items.append(new_line_item)
        
    # done with grouping, update Qty values for each list items
    for idx, li in enumerate(line_items):
        li['Qty'] = len(li['refdes_list'])
        li['item'] = idx
        
    return line_items  

def build_bom(output_file, parts_database):
    """ build BOM """
    pass 

if __name__ == "__main__":

    # open xml file and pull parts database from it.
    xmlfile = sys.argv[1]
    xlsfile = sys.argv[2]
    print(f"Read xml source data file: {xmlfile}")
    fh = open(xmlfile, "r")
    xml_content = fh.read()
    ordered_dict = xmltodict.parse(xml_content)
    parts_database = ordered_dict['export']['components']['comp']
    if check_bom(parts_database) != True:
        print("BOM Check Failed..")
        sys.exit()
        
    print("BOM Check passes..")
    line_items = group_bom_items(parts_database)
    if line_items == -1: 
        sys.exit(-1)
        
    print("Line Items list:")
    idx = 1
    for li in line_items:
        print(f" processing:  {str(li)} ")
            
        
    print(f"Found {idx-1} line items from {len(parts_database)} components in BOM database.")
    
    # write this to excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # create column headers
    headers = [ x for x in line_items[0] ]   
    
    for index, value in enumerate(headers):
        sheet.cell(row=1, column=index+1).value = value
            
    for i, li in enumerate(line_items):
        for idx,value in enumerate(li.values()):
            if type(value ) == list:
                sheet.cell(row=i+2, column=idx+1).value = ', '.join(value)
            else:
                sheet.cell(row=i+2, column=idx+1).value = value       
            
    workbook.save(xlsfile)
    print(f"Wrote output to {xlsfile}")
    print("Done..")
    
    
    
    